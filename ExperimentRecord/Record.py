import pandas as pd
import os
import sys
from typing import Dict
from openpyxl import load_workbook
from itertools import zip_longest


def ParaSettingRecord(target_dir, algorithm_name, to_write: dict):
    # create a corresponding excel file if not exist
    if not os.path.exists(os.path.join(target_dir, algorithm_name + "_parasetting.csv")):
        df = pd.DataFrame(to_write)
        df.to_csv(os.path.join(target_dir, algorithm_name + "_parasetting.csv"), header=True, index=None)

    else:
        # record each experiment's parameters setting
        df = pd.DataFrame(to_write, columns=list(to_write.keys()))
        df.to_csv(os.path.join(target_dir, algorithm_name + "_parasetting.csv"), mode='a', header=False,
                  index=None)


def ResultRecord(target_dir, algorithm_name, to_write: dict):
    if not os.path.exists(os.path.join(target_dir, algorithm_name + "_result_record.xlsx")):
        df = pd.DataFrame(columns=list(to_write.keys()))
        df.to_excel(os.path.join(target_dir, algorithm_name + "_result_record.xlsx"), sheet_name=to_write["实验编号"][0],
                    index=False)

    values = [value for value in list(to_write.values())]
    df = pd.DataFrame(
        zip_longest(values[0], values[1], values[2], values[3], values[4], values[5], values[6], values[7], values[8],
                    values[9]))

    book = load_workbook(os.path.join(target_dir, algorithm_name + "_result_record.xlsx"))

    if to_write["实验编号"][0] in book.sheetnames:  # if write mode="a"
        df1 = pd.DataFrame(
            pd.read_excel(os.path.join(target_dir, algorithm_name + "_result_record.xlsx"),
                          sheet_name=to_write["实验编号"][0]))
        writer = pd.ExcelWriter(os.path.join(target_dir, algorithm_name + "_result_record.xlsx"), engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_rows = df1.shape[0]
        df.to_excel(writer, sheet_name=to_write["实验编号"][0], startrow=df_rows + 1, index=False,
                    header=False)
        writer.save()
    elif to_write["实验编号"][0] not in book.sheetnames:
        writer = pd.ExcelWriter(os.path.join(target_dir, algorithm_name + "_result_record.xlsx"), engine='openpyxl')
        writer.book = book
        header = pd.DataFrame(
            zip_longest(values[0], values[1], values[2], values[3], values[4], values[5], values[6], values[7],
                        values[8], values[9]), columns=list(to_write.keys()))
        header.to_excel(writer, sheet_name=to_write["实验编号"][0],
                        header=True, index=None)
        writer.save()
        writer.close()
