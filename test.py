'''
t = list(range(10))
print(t)
zero = [0]
zero.extend(t)
print(zero)

import numpy as np

a = np.array(range(10))
print(type(a))
print(sum(a))
print(sum(t))
'''

import random
import numpy as np


def dis2one(target):
    if abs(target - 1) < 1e-4:
        return True
    else:
        return False


def RandomWeight():
    while 1:
        wa = random.random()
        wb = random.random()
        wd = random.random()
        print("Respectivelu: {:.4f}, {:.4f}, {:.4f}".format(wa, wb, wd))
        if wa > wb and wb > wd and wa > wd and dis2one((wa+wb+wd)):
            return np.array([wa, wb, wd])


RandomWeight()

from itertools import zip_longest
a=[1,2,3,4]
b=["a","b","c","d"]
c=[5,6,7,8]
l=[a,b,c]



import pandas as pd

#df=pd.DataFrame(zip_longest(a,b,c),columns=["a","b","c"])
# df.to_excel("test.xlsx",sheet_name="123",index=None,header=True)

to_write={"a":1,"b":2,"c":3}
print(list(to_write.values()))

a={"PSNR":["5"]}
def change(a):
    a["PSNR"].append(str(6))
    return a
a=change(a)
print(a)

from openpyxl import load_workbook
book=load_workbook(r"/Users/moka/Desktop/ExpRunningResult/Tsallis_Entropy/EGWO_result_record.xlsx")
print(book.sheetnames)

df=pd.DataFrame(zip_longest(a,b,c))
print(df)
header=pd.DataFrame(zip_longest(a,b,c),columns=['algorithm','threshold','PSNR'])
print(header)